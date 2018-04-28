import openpyxl
import sys
import os

path = os.path.dirname(sys.argv[0])
file = sys.argv[3]

wb = openpyxl.load_workbook(os.path.join(path,file))
sheet = wb.active

wb_new = openpyxl.Workbook()
sheet_new = wb_new.active


row = int(sys.argv[1])
blank_rows = int(sys.argv[2])

column = openpyxl.utils.get_column_letter(sheet.max_column)
top_area = sheet['A1:' + column + str(row)]
bottom_area = sheet['A' + str(row + 1) + ':' + column + str(sheet.max_row)]


for cellObj in top_area:
    for cell in cellObj:
        coordinates = cell.coordinate
        sheet_new[coordinates]=cell.value
        
for cellObj in bottom_area:
    for cell in cellObj:
        letter = cell.column
        num = cell.row
        sheet_new[letter+str(num+blank_rows)]=cell.value

wb_new.save(r'YOUR_PATH\BlankRowsInserter.xlsx')
