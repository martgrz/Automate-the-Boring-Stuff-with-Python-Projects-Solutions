#   openpyxl.__version__
#   '2.5.3'

import openpyxl

wb = openpyxl.load_workbook('your_file_name.xlsx')
sheet = wb.active

wb_new = openpyxl.Workbook()
sheet_new = wb_new.active

max_row = sheet.max_row
max_column = openpyxl.utils.get_column_letter(sheet.max_column)

cells = sheet['A1:' + max_column + str(max_row)]


for cellObj in cells:
    for cell in cellObj:
        
        x = cell.row
        y = openpyxl.utils.column_index_from_string(cell.column)

        sheet_new.cell(y,x).value = sheet.cell(x,y).value


wb_new.save('cell_inverter.xlsx')

        

