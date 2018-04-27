#   openpyxl.__version__
#   '2.5.3'

import openpyxl
import sys
from openpyxl.styles import Font

num = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb['Sheet']

for i in range(num):
    sheet.cell(2+i,1).value = i+1
    sheet.cell(2+i,1).font = Font(bold=True)
    sheet.cell(1,2+i).value = i+1
    sheet.cell(1,2+i).font = Font(bold=True)

column = openpyxl.utils.get_column_letter(num+1)

cells_range = sheet[('b2:'+column+str(num+1))]

    
for cellObj in cells_range:
    for cell in cellObj:
        cell.value=(sheet[cell.column + '1'].value*sheet['A'+str(cell.row)].value)

wb.save(r'YOUR_PATH\multiplicationTable.xlsx')
